import numpy as np
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# LECTURA DEL ARCHIVO
# ---------------------------------------------------------------------------

def _find_matrix_bounds(ws):
    """
    Detecta automáticamente la fila de encabezado de la tabla MICMAC.
    El encabezado tiene col A = None y col B = nombre de primera variable.
    La fila siguiente tiene col A = ese mismo nombre (diagonal).
    """
    for r in range(1, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        if a_val is None and isinstance(b_val, str) and b_val.strip():
            next_a = ws.cell(row=r + 1, column=1).value
            if isinstance(next_a, str) and next_a.strip() == b_val.strip():
                header_row = r
                variables = []
                col = 2
                while True:
                    val = ws.cell(row=header_row, column=col).value
                    if val is None:
                        break
                    variables.append(str(val).strip())
                    col += 1
                return header_row + 1, len(variables), variables

    raise ValueError("No se encontró la tabla MICMAC en la hoja MATRIZ.")


def _extract_matrix(ws, data_start_row, n):
    """Extrae la matriz numérica n x n."""
    M = np.zeros((n, n), dtype=float)
    for i in range(n):
        for j in range(n):
            val = ws.cell(row=data_start_row + i, column=2 + j).value
            if isinstance(val, (int, float)):
                M[i][j] = float(val)
    return M


def _load_descriptors(wb):
    """Mapeo NOMBRE_CORTO -> DESCRIPTOR desde hoja DESCRIPTORES."""
    sheet_name = next(
        (s for s in wb.sheetnames if s.strip().upper() == "DESCRIPTORES"), None
    )
    if sheet_name is None:
        raise ValueError("No se encontró la hoja DESCRIPTORES.")
    ws = wb[sheet_name]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        corto, descriptor = row[0], row[1]
        if corto and descriptor:
            mapping[str(corto).strip()] = str(descriptor).strip()
    return mapping


# ---------------------------------------------------------------------------
# ALGORITMO LIPSOR OFICIAL
# ---------------------------------------------------------------------------

def _ranking(values):
    """Ranking descendente (0 = mayor valor)."""
    order = np.argsort(-values)
    ranks = np.empty_like(order)
    ranks[order] = np.arange(len(values))
    return ranks


def _find_stable_k(M, k_min=2, k_max=8):
    """
    Busca el k minimo donde el ranking de motricidad Y dependencia
    se estabiliza entre k y k+1.
    Replica el criterio de estabilidad del software LIPSOR oficial.
    """
    prev_rank_I = None
    prev_rank_D = None

    for k in range(k_min, k_max + 1):
        Mk = np.linalg.matrix_power(M, k)
        rank_I = _ranking(Mk.sum(axis=1))
        rank_D = _ranking(Mk.sum(axis=0))

        if prev_rank_I is not None:
            if np.array_equal(rank_I, prev_rank_I) and np.array_equal(rank_D, prev_rank_D):
                return k - 1  # el anterior ya era el estable

        prev_rank_I = rank_I
        prev_rank_D = rank_D

    return k_max  # si no convergió, usar el máximo


def _classify_micmac(variables, M):
    """
    Algoritmo LIPSOR oficial:
      1. MII = M^k  (potencia pura hasta estabilidad del ranking)
      2. Motricidad = suma filas MII
      3. Dependencia = suma columnas MII
      4. Umbral = promedio de cada eje (criterio LIPSOR)
      5. Clasificar en cuadrantes
    """
    k = _find_stable_k(M)
    MII = np.linalg.matrix_power(M, k)

    motricidad  = MII.sum(axis=1)
    dependencia = MII.sum(axis=0)

    umbral_I = motricidad.mean()
    umbral_D = dependencia.mean()

    cuadrantes = {"poder": [], "conflicto": [], "resultados": [], "autonomas": []}

    for i, var in enumerate(variables):
        alta_I = motricidad[i] >= umbral_I
        alta_D = dependencia[i] >= umbral_D

        if alta_I and not alta_D:
            cuadrantes["poder"].append(var)
        elif alta_I and alta_D:
            cuadrantes["conflicto"].append(var)
        elif not alta_I and alta_D:
            cuadrantes["resultados"].append(var)
        else:
            cuadrantes["autonomas"].append(var)

    return cuadrantes, k


# ---------------------------------------------------------------------------
# ESCRITURA EN INFO_ENGINE
# ---------------------------------------------------------------------------

def _write_to_engine(wb_engine, cuadrantes, descriptor_map):
    """
    Escribe resultados en info_engine a partir de fila 124:
      B = Poder | C = Conflicto | D = Resultados | E = Autonomas
    Una variable por fila bajando desde 124.
    Usa DESCRIPTOR completo en lugar del nombre corto.
    """
    col_map = {
        "poder":      2,   # B
        "conflicto":  3,   # C
        "resultados": 4,   # D
        "autonomas":  5,   # E
    }
    sheet = wb_engine.active
    start_row = 124

    for cuadrante, col in col_map.items():
        for offset, var_corto in enumerate(cuadrantes[cuadrante]):
            descriptor = descriptor_map.get(var_corto, var_corto)
            sheet.cell(row=start_row + offset, column=col, value=descriptor)


# ---------------------------------------------------------------------------
# FUNCIÓN PRINCIPAL (llamada desde app.py)
# ---------------------------------------------------------------------------

def procesar_micmac(archivo_micmac, wb_engine):
    """
    archivo_micmac : file-like object (BytesIO desde Streamlit)
    wb_engine      : workbook openpyxl de info_engine ya cargado
    """
    wb = load_workbook(archivo_micmac)

    descriptor_map = _load_descriptors(wb)

    ws_matriz = wb["MATRIZ"]
    data_start_row, n, variables = _find_matrix_bounds(ws_matriz)
    M = _extract_matrix(ws_matriz, data_start_row, n)

    cuadrantes, k_usado = _classify_micmac(variables, M)

    _write_to_engine(wb_engine, cuadrantes, descriptor_map)

    return cuadrantes, k_usado
