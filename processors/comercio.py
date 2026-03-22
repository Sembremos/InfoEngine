import pandas as pd
from openpyxl import load_workbook
import io


def contar_frecuencias(df, columna, orden):
    serie = df[columna].dropna().astype(str).str.strip().str.lower()
    conteo = serie.value_counts()
    return [int(conteo.get(opcion, 0)) for opcion in orden]


def contar_multiple(df, columna, orden):
    serie = df[columna].dropna().astype(str)

    todas = []
    for valor in serie:
        partes = [x.strip().lower() for x in valor.split(",")]
        todas.extend(partes)

    conteo = pd.Series(todas).value_counts()
    return [int(conteo.get(opcion, 0)) for opcion in orden]


def escribir_lista(ws, columna, fila_inicio, lista):
    for i, valor in enumerate(lista):
        try:
            ws[f"{columna}{fila_inicio+i}"] = int(valor)
        except:
            ws[f"{columna}{fila_inicio+i}"] = 0


def procesar_comercio(df, wb): 
    ws = wb["Hoja1"]

    # 1 DISTRITOS
    # leer distritos desde Excel (columna A)
    distritos_excel = []
    
    for i in range(16):
        valor = ws[f"A{8+i}"].value
        if valor:
            distritos_excel.append(str(valor).strip().lower())
        else:
            distritos_excel.append(None)
    
    # conteo desde dataframe
    conteo = df["2. Distrito:"].dropna().astype(str).str.strip().str.lower().value_counts()
    
    # construir frecuencias respetando el orden del Excel
    frec = []
    
    for d in distritos_excel:
        if d is None:
            frec.append(0)
        else:
            frec.append(int(conteo.get(d, 0)))
    
    escribir_lista(ws, "D", 8, frec)

    # 2 EDAD
    orden = [
        "18_a_29_anos",
        "30_a_44_anos",
        "45_a_64_anos",
        "65_anos_o_mas",
        "vacio"
    ]

    serie = df["3. Edad (en años cumplidos): marque una categoría que incluya su edad."]
    serie_limpia = serie.dropna().astype(str).str.strip().str.lower()
    conteo = serie_limpia.value_counts()

    frec = []
    for op in orden:
        if op == "vacio":
            frec.append(serie.isna().sum())
        else:
            frec.append(conteo.get(op, 0))

    escribir_lista(ws, "E", 29, frec)

    # 3 ESCOLARIDAD
    orden = [
        "ninguna",
        "primaria_completa",
        "primaria_incompleta",
        "secundaria_completa",
        "secundaria_incompleta",
        "tecnico",
        "universitaria_completa",
        "universitaria_incompleta"
    ]

    frec = contar_frecuencias(df, "5. Escolaridad:", orden)
    escribir_lista(ws, "E", 39, frec)

    # 4 GENERO
    orden = ["masculino", "femenino", "persona_no_binaria"]
    frec = contar_frecuencias(df, "4. ¿Con cuál de estas opciones se identifica?", orden)
    escribir_lista(ws, "E", 52, frec)

    # 5 VICTIMA
    orden = ["no", "si_pero_no_denuncie", "si_y_denuncie"]
    frec = contar_frecuencias(df,
        "23. Durante los últimos 12 meses, ¿su local comercial fue afectado por algún delito?",
        orden
    )
    escribir_lista(ws, "E", 314, frec)

    # 6 NO DENUNCIA (multiple)
    orden = [
        "distancia_o_dificultad_de_acceso_a_oficinas_para_denunciar",
        "miedo_a_represalias",
        "falta_de_respuesta_o_seguimiento_en_denuncias_anteriores",
        "complejidad_al_colocar_la_denuncia",
        "desconocimiento_de_donde_colocar_la_denuncia",
        "el_policia_me_dijo_que_era_mejor_no_denunciar",
        "falta_de_tiempo_para_colocar_la_denuncia",
        "desconfianza_en_las_autoridades_o_en_el_proceso_de_denuncia",
        "otro_motivo"
    ]

    frec = contar_multiple(df,
        "23.2 En caso de NO haber realizado la denuncia ante el OIJ, indique cuál fue el motivo:",
        orden
    )
    escribir_lista(ws, "E", 322, frec)

    # 7 HORARIO (multiple)
    orden = [
        "00_00_02_59_madrugada",
        "03_00_05_59_madrugada",
        "06_00_08_59_manana",
        "09_00_11_59_manana",
        "12_00_14_59_mediodia_tarde",
        "15_00_17_59_tarde",
        "18_00_20_59_noche",
        "21_00_23_59_noche",
        "desconocido"
    ]

    frec = contar_multiple(df,
        "23.3 ¿Tiene conocimiento del horario en el cual se presentó el hecho delictivo que afectó a su local comercial o a personas vinculadas a su actividad comercial?",
        orden
    )
    escribir_lista(ws, "E", 336, frec)

    # 8 METODO (multiple)
    orden = [
        "arma_blanca_cuchillo_machete_tijeras",
        "arma_de_fuego",
        "amenazas",
        "arrebato",
        "boquete",
        "ganzua_pata_de_chancho",
        "engano",
        "escalamiento",
        "no_se",
        "otro"
    ]

    frec = contar_multiple(df,
        "24. ¿Cuál fue la forma o modo en que ocurrió la situación que afectó a su local comercial?",
        orden
    )
    escribir_lista(ws, "E", 350, frec)

    # 9 SERVICIO
    orden = ["peor_servicio", "igual", "mejor_servicio"]
    frec = contar_frecuencias(df,
        "27. ¿Cómo ha sido el servicio policial de Fuerza Pública de Costa Rica en los últimos 24 meses?",
        orden
    )
    escribir_lista(ws, "E", 373, frec)

    # 10 CONOCE FP
    orden = ["no", "si"]
    frec = contar_frecuencias(df,
        "28. ¿Conoce usted a los policías de la Fuerza Pública de Costa Rica de su zona comercial?",
        orden
    )
    escribir_lista(ws, "E", 381, frec)

    # 11 SEGURIDAD
    orden = ["muy_inseguro", "inseguro", "ni_seguro_ni_inseguro", "seguro", "muy_seguro"]
    frec = contar_frecuencias(df,
        "7. ¿Qué tan seguro percibe usted el entorno en su local comercial?",
        orden
    )
    escribir_lista(ws, "D", 398, frec)

    # 12 CONOCE PROGRAMA
    frec = contar_frecuencias(df,
        '29. ¿Conoce el programa de "Seguridad Comercial" que imparte Fuerza Pública?',
        ["no", "si"]
    )
    escribir_lista(ws, "D", 406, frec)

    # 13 INSCRITO
    frec = contar_frecuencias(df,
        '30. ¿Está inscrito en el programa de "Seguridad Comercial" que imparte Fuerza Pública?',
        ["no", "si"]
    )
    escribir_lista(ws, "D", 413, frec)

    # 14 CONTACTO
    frec = contar_frecuencias(df,
        "31. ¿Le gustaría que se le contacte para formar parte del programa?",
        ["no", "si"]
    )
    escribir_lista(ws, "D", 420, frec)

    archivo = io.BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    return archivo
