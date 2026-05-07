from openpyxl import load_workbook


def procesar_lineas_accion(archivo_lineas, wb_info):

    wb = load_workbook(archivo_lineas, data_only=True)
    ws_info = wb_info.active

    # -----------------------------
    # CONFIGURACIONES
    columnas_objetivos = [
        7, 13, 19, 25, 31, 37, 43, 49, 55, 61, 67, 73
    ]  # G, M, S, Y...

    columnas_tablas = [
        (9, 10), (15, 16), (21, 22), (27, 28),
        (33, 34), (39, 40), (45, 46), (51, 52),
        (57, 58), (63, 64), (69, 70), (75, 76)
    ]

    columnas_cogestores = [
        9, 15, 21, 27, 33, 39,
        45, 51, 57, 63, 69, 75
    ]

    # -----------------------------
    # CONTADORES
    conteo_fp = 0
    conteo_muni = 0
    conteo_mixto = 0

    # -----------------------------
    # RECORRER HOJAS
    hojas = [h for h in wb.sheetnames if "Línea de Acción" in h]

    for idx, nombre in enumerate(hojas):

        ws = wb[nombre]

        # -----------------------------
        # 1. TIPO DE LINEA
        lideres = []
        fila = 5

        while True:
            valor = ws.cell(row=fila, column=6).value  # F
            accion = ws.cell(row=fila, column=2).value  # B

            if not accion:
                break

            if valor:
                lideres.append(str(valor).strip())

            fila += 1

        set_lideres = set(lideres)

        if set_lideres == {"Fuerza Pública"}:
            conteo_fp += 1
        elif set_lideres == {"Municipalidad"}:
            conteo_muni += 1
        else:
            conteo_mixto += 1

        # -----------------------------
        # 2. OBJETIVOS
        if idx < len(columnas_objetivos):
            col = columnas_objetivos[idx]
            ws_info.cell(row=245, column=col, value=ws["C3"].value)

        # -----------------------------
        # 3. PROBLEMATICAS
        texto = ws["C2"].value
        if texto:
            partes = [p.strip() for p in str(texto).split(",")]

            fila_dest = 242 + idx

            for i in range(3):
                valor = partes[i] if i < len(partes) else None
                ws_info.cell(row=fila_dest, column=2 + i, value=valor)

        # -----------------------------
        # 4. ACCIONES + LIDER
        if idx < len(columnas_tablas):
            col_acc, col_lid = columnas_tablas[idx]

            fila = 5
            fila_dest = 249

            while fila_dest <= 257:
                accion = ws.cell(row=fila, column=2).value
                lider = ws.cell(row=fila, column=6).value

                if not accion:
                    break

                # transformar lider
                if lider == "Fuerza Pública":
                    lider_txt = "FP"
                elif lider == "Municipalidad":
                    lider_txt = "GL"
                else:
                    lider_txt = None

                ws_info.cell(row=fila_dest, column=col_acc, value=accion)
                ws_info.cell(row=fila_dest, column=col_lid, value=lider_txt)

                fila += 1
                fila_dest += 1

        # -----------------------------
        # 5. COGESTORES
        if idx < len(columnas_cogestores):
            col = columnas_cogestores[idx]
        
            fila = 5
            lista = []
        
            while True:
                accion = ws.cell(row=fila, column=2).value
                cog = ws.cell(row=fila, column=7).value
        
                if not accion:
                    break
        
                if cog:
                    cog_limpio = str(cog).strip()
        
                    # evitar duplicados manteniendo orden
                    if cog_limpio not in lista:
                        lista.append(cog_limpio)
        
                fila += 1
        
            texto_final = ", ".join(lista)
            ws_info.cell(row=262, column=col, value=texto_final)

    # -----------------------------
    # GUARDAR CONTEOS
    ws_info["B239"] = conteo_fp
    ws_info["A239"] = conteo_muni
    ws_info["C239"] = conteo_mixto
