import pandas as pd
from openpyxl import load_workbook


# -----------------------------
# PROCESAR PARETO
# -----------------------------
def procesar_pareto(archivo_pareto, wb_destino):

    df = pd.read_excel(archivo_pareto)
    df_desglose = pd.read_excel(archivo_pareto, sheet_name="Desglose")
    df.columns = [str(c).strip().upper() for c in df.columns]

    col_categoria = "CATEGORIA"
    col_descriptor = "DESCRIPTOR"
    col_segmento = "SEGMENTO"

    if col_categoria not in df.columns:
        col_categoria = df.columns[0]
    if col_descriptor not in df.columns:
        col_descriptor = df.columns[1]
    if col_segmento not in df.columns:
        col_segmento = df.columns[6]

    # -----------------------------
    # FILTRAR PRIORIZADOS
    # -----------------------------
    df_filtrado = df[df[col_segmento].astype(str).str.upper() == "PRIORIZADO"]

    # -----------------------------
    # HOJAS
    # -----------------------------
    ws_hoja1 = wb_destino["Hoja1"]
    ws_pareto = wb_destino["pareto"] 

    # -----------------------------
    # TABLA COMPLETA → HOJA PARETO
    # -----------------------------
    fila_excel = 2

    for _, row in df_filtrado.iterrows():

        if fila_excel > 31:
            break

        ws_pareto[f"A{fila_excel}"] = row.iloc[0]
        ws_pareto[f"B{fila_excel}"] = row.iloc[1]
        ws_pareto[f"C{fila_excel}"] = row.iloc[2]
        ws_pareto[f"D{fila_excel}"] = row.iloc[3]
        ws_pareto[f"E{fila_excel}"] = row.iloc[4]
        ws_pareto[f"F{fila_excel}"] = row.iloc[5]
        ws_pareto[f"G{fila_excel}"] = row.iloc[6]

        fila_excel += 1

    # -----------------------------
    # TOTAL → HOJA1
    # -----------------------------
    df_raw = pd.read_excel(archivo_pareto, header=None)

    total_valor = None

    for i in range(len(df_raw)):
        celda = str(df_raw.iloc[i, 1]).strip().upper()

        if celda == "TOTAL:":
            total_valor = df_raw.iloc[i, 2]
            break

    if total_valor is not None:
        ws_hoja1["B88"] = total_valor

    # -----------------------------
    # CANTIDAD DESCRIPTORES → HOJA1
    # -----------------------------
    ws_hoja1["D93"] = df[col_descriptor].dropna().shape[0]

    # -----------------------------
    # DELITOS / RIESGOS → HOJA1
    # -----------------------------
    delitos = df_filtrado[df_filtrado[col_categoria].astype(str).str.upper() == "DELITO"]
    riesgos = df_filtrado[df_filtrado[col_categoria].astype(str).str.upper() == "RIESGO SOCIAL"]

    lista_delitos = delitos[col_descriptor].tolist()
    lista_riesgos = riesgos[col_descriptor].tolist()

    for fila in range(97, 118):
        ws_hoja1[f"B{fila}"] = None
        ws_hoja1[f"C{fila}"] = None

    fila = 97
    for item in lista_delitos:
        if fila > 117:
            break
        ws_hoja1[f"B{fila}"] = item
        fila += 1

    fila = 97
    for item in lista_riesgos:
        if fila > 117:
            break
        ws_hoja1[f"C{fila}"] = item
        fila += 1
    # -----------------------------
    # NUEVAS FUNCIONES (DESGLOSE)
    # -----------------------------
    extraer_totales_desglose(df_desglose, ws_hoja1)
    escribir_frecuencias_problematicas(df_desglose, ws_hoja1)

#____________________ F2---- extraccion

def extraer_totales_desglose(df_desglose, ws_hoja1):

    # normalizar columnas
    columnas = {c.upper(): c for c in df_desglose.columns}

    def buscar_columna(nombre_base):
        for col_upper, col_real in columnas.items():
            if nombre_base in col_upper:
                return col_real
        return None

    mapa = {
        "PARETO COMUNIDAD": "B83",
        "PARETO COMERCIO": "B84",
        "PARETO POLICIAL": "B85",
        "PARETO OPO": "B86",
        "PARETO SAE": "B87"
    }

    for clave, celda in mapa.items():
        col = buscar_columna(clave)

        if col:
            total = pd.to_numeric(df_desglose[col], errors="coerce").fillna(0).sum()
            ws_hoja1[celda] = total
        else:
            ws_hoja1[celda] = 0


#f3,,,,,, desglose

def escribir_frecuencias_problematicas(df_desglose, ws_hoja1):

    # normalizar nombres
    df_desglose.columns = [str(c).strip().upper() for c in df_desglose.columns]

    # detectar columnas dinámicas
    def buscar_columna(nombre):
        for col in df_desglose.columns:
            if nombre in col:
                return col
        return None

    col_descriptor = buscar_columna("DESCRIPTOR")
    col_comunidad = buscar_columna("PARETO COMUNIDAD")
    col_comercio = buscar_columna("PARETO COMERCIO")

    if not col_descriptor:
        return

    # columnas destino
    columnas_comunidad = [78, 80, 82]
    columnas_comercio = [79, 81, 83]

    for fila in range(242, 254):

        # limpiar primero
        for col in columnas_comunidad + columnas_comercio:
            ws_hoja1.cell(row=fila, column=col, value=None)

        # leer problemáticas (B, C, D)
        problemas = [
            ws_hoja1.cell(row=fila, column=2).value,
            ws_hoja1.cell(row=fila, column=3).value,
            ws_hoja1.cell(row=fila, column=4).value
        ]

        for idx, problema in enumerate(problemas):

            if not problema:
                continue

            filtro = df_desglose[
                df_desglose[col_descriptor].astype(str).str.strip().str.upper() ==
                str(problema).strip().upper()
            ]

            if filtro.empty:
                continue

            fila_df = filtro.iloc[0]

            # comunidad
            if col_comunidad:
                ws_hoja1.cell(
                    row=fila,
                    column=columnas_comunidad[idx],
                    value=fila_df[col_comunidad]
                )

            # comercio
            if col_comercio:
                ws_hoja1.cell(
                    row=fila,
                    column=columnas_comercio[idx],
                    value=fila_df[col_comercio]
                )


  
