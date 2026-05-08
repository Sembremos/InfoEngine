from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



# ---------------------------------------------------------
# NORMALIZAR TEXTO
# ---------------------------------------------------------
import re
import unicodedata


# ---------------------------------------------------------
# NORMALIZAR TEXTO INTELIGENTE
# ---------------------------------------------------------
def normalizar_texto(texto):

    if not texto:
        return ""

    texto = str(texto).lower().strip()

    # quitar tildes
    texto = ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )

    # eliminar contenido entre parentesis
    texto = re.sub(r'\(.*?\)', ' ', texto)

    # eliminar caracteres raros
    texto = re.sub(r'[^a-z0-9\s]', ' ', texto)

    # palabras irrelevantes
    stopwords = {
        "de",
        "del",
        "la",
        "las",
        "el",
        "los",
        "en",
        "y",
        "situacion",
        "publica",
        "publico"
    }

    palabras = []

    for palabra in texto.split():

        # singular simple
        if palabra.endswith("s") and len(palabra) > 4:
            palabra = palabra[:-1]

        if palabra not in stopwords:
            palabras.append(palabra)

    return " ".join(palabras)


# ---------------------------------------------------------
# COMPARAR PROBLEMÁTICAS
# ---------------------------------------------------------
def son_similares(texto1, texto2):

    t1 = normalizar_texto(texto1)
    t2 = normalizar_texto(texto2)

    if not t1 or not t2:
        return False

    # exacta
    if t1 == t2:
        return True

    # parcial
    if t1 in t2 or t2 in t1:
        return True

    palabras1 = set(t1.split())
    palabras2 = set(t2.split())

    interseccion = palabras1.intersection(palabras2)

    # si comparten suficientes palabras
    if len(interseccion) >= 2:
        return True

    return False


# ---------------------------------------------------------
# LIMPIAR LISTA
# ELIMINA DUPLICADOS Y VACÍOS
# ---------------------------------------------------------
def limpiar_lista(lista):
    resultado = []
    vistos = set()

    for item in lista:

        if not item:
            continue

        texto = str(item).strip()

        if not texto:
            continue

        texto_normalizado = texto.lower()

        if texto_normalizado not in vistos:
            vistos.add(texto_normalizado)
            resultado.append(texto)

    return resultado


# ---------------------------------------------------------
# OBTENER CAUSAS DESDE UNA HOJA
# ---------------------------------------------------------
def obtener_causas_hoja(hoja):

    fila_encabezados = 4

    col_socio = None
    col_estructural = None

    # Buscar columnas
    for col in range(1, hoja.max_column + 1):

        valor = hoja.cell(row=fila_encabezados, column=col).value

        if valor:

            valor = str(valor).strip().lower()

            if "socio" in valor:
                col_socio = col

            elif "estructural" in valor:
                col_estructural = col

    if not col_socio or not col_estructural:
        return [], []

    socio = []
    estructural = []

    fila = 5

    while True:

        val_socio = hoja.cell(row=fila, column=col_socio).value
        val_estructural = hoja.cell(row=fila, column=col_estructural).value

        # FIN TABLA
        if not val_socio and not val_estructural:
            break

        if val_socio:
            socio.append(str(val_socio).strip())

        if val_estructural:
            estructural.append(str(val_estructural).strip())

        fila += 1

    return limpiar_lista(socio), limpiar_lista(estructural)


# ---------------------------------------------------------
# FUNCIÓN PRINCIPAL
# ---------------------------------------------------------
def procesar_triangulo(archivo_triangulo, wb_destino):

    wb_origen = load_workbook(archivo_triangulo, data_only=True)

    hoja_destino = wb_destino["Hoja1"]

    # ---------------------------------------------------------
    # PARTE ORIGINAL
    # CONTADORES
    # ---------------------------------------------------------
    total_socio = 0
    total_estructural = 0

    datos_hojas = {}

    # ---------------------------------------------------------
    # LEER TODAS LAS HOJAS
    # ---------------------------------------------------------
    for hoja in wb_origen.worksheets:

        nombre_hoja = normalizar_texto(hoja.title)

        socio, estructural = obtener_causas_hoja(hoja)

        total_socio += len(socio)
        total_estructural += len(estructural)

        datos_hojas[nombre_hoja] = {
            "socio": socio,
            "estructural": estructural
        }

    # Escribir contadores originales
    hoja_destino["B147"] = total_socio
    hoja_destino["C147"] = total_estructural

    # ---------------------------------------------------------
    # MAPEO FILAS -> COLUMNAS DESTINO
    # ---------------------------------------------------------
    mapa_columnas = {
        242: "F",
        243: "L",
        244: "R",
        245: "X",
        246: "AD",
        247: "AJ",
        248: "AP",
        249: "AV",
        250: "BB",
        251: "BH",
        252: "BN",
        253: "BT"
    }

    # ---------------------------------------------------------
    # RECORRER LÍNEAS DE ACCIÓN
    # ---------------------------------------------------------
    for fila in range(242, 254):

        columna_destino = mapa_columnas[fila]

        causas_socio = []
        causas_estructural = []

        # ---------------------------------------------------------
        # PROBLEMÁTICAS B-C-D
        # ---------------------------------------------------------
        for col in ["B", "C", "D"]:

            valor = hoja_destino[f"{col}{fila}"].value

            if not valor:
                continue

            problema = valor

            for nombre_hoja, datos in datos_hojas.items():
            
                if son_similares(problema, nombre_hoja):
            
                    causas_socio.extend(
                        datos["socio"]
                    )
            
                    causas_estructural.extend(
                        datos["estructural"]
                    )

        # ---------------------------------------------------------
        # LIMPIAR DUPLICADOS
        # ---------------------------------------------------------
        causas_socio = limpiar_lista(causas_socio)
        causas_estructural = limpiar_lista(causas_estructural)

        # ---------------------------------------------------------
        # UNIR LISTAS
        # Primero socio
        # Luego estructurales
        # ---------------------------------------------------------
        lista_final = []

        lista_final.extend(causas_socio)
        lista_final.extend(causas_estructural)

        # Máximo espacio disponible
        lista_final = lista_final[:30]

        # ---------------------------------------------------------
        # ESCRIBIR
        # F247:F276 etc
        # ---------------------------------------------------------
        fila_inicio = 247

        for i, causa in enumerate(lista_final):

            fila_destino = fila_inicio + i

            hoja_destino[
                f"{columna_destino}{fila_destino}"
            ] = causa
