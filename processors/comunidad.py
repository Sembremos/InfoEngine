import pandas as pd
from openpyxl import load_workbook
import io


# --------------------------------------
# FUNCIONES GENERALES
# --------------------------------------

def contar_frecuencias(df, columna, orden):
    serie = df[columna]

    # normalizar datos (sin eliminar vacíos)
    serie_limpia = serie.dropna().astype(str).str.strip().str.lower()
    conteo = serie_limpia.value_counts()

    resultado = []

    for opcion in orden:
        if opcion.strip().lower() == "vacio":
            resultado.append(serie.isna().sum())
        else:
            resultado.append(conteo.get(opcion.strip().lower(), 0))

    return resultado


def escribir_lista(ws, columna, fila_inicio, lista):
    for i, valor in enumerate(lista):
        try:
            ws[f"{columna}{fila_inicio+i}"] = int(valor)
        except:
            ws[f"{columna}{fila_inicio+i}"] = 0


def limpiar_lista(ws, columna, fila_inicio, cantidad):
    for i in range(cantidad):
        ws[f"{columna}{fila_inicio+i}"] = None


def formatear_canton(texto):
    texto = str(texto).replace("_", " ").title()
    texto = texto.replace(" Ramon", " Ramón")
    return texto


# --------------------------------------
# FUNCION PRINCIPAL
# --------------------------------------

def procesar_comunidad(archivo_comunidad):

    df = pd.read_excel(archivo_comunidad)

    import streamlit as st
    st.write(df.columns.tolist())

    wb = load_workbook("plantillas/info_engine.xlsx")
    ws = wb["Hoja1"]

    # -----------------------------------
    # 1 CANTON
    # -----------------------------------

    canton_raw = df["1. Cantón:"].dropna().iloc[0]
    ws["B2"] = formatear_canton(canton_raw)

    # -----------------------------------
    # 2 LISTA DISTRITOS
    # -----------------------------------
    
    distritos = sorted(df["2. Distrito:"].dropna().unique())
    
    # limitar a máximo 16
    distritos = distritos[:16]
    
    # escribir distritos
    for i in range(16):
        if i < len(distritos):
            ws[f"A{8+i}"] = distritos[i]
        else:
            ws[f"A{8+i}"] = None
    
    # conteo
    conteo = df["2. Distrito:"].value_counts()
    
    frecuencias = []
    for d in distritos:
        frecuencias.append(conteo.get(d, 0))
    
    # completar hasta 16
    while len(frecuencias) < 16:
        frecuencias.append(0)
    
    escribir_lista(ws, "E", 8, frecuencias)

    # -----------------------------------
    # 4 RELACION ZONA
    # -----------------------------------

    orden = [
        "estudio_en_la_zona",
        "trabajo_en_la_zona",
        "visito_la_zona",
        "vivo_en_la_zona"
    ]

    frec = contar_frecuencias(df, "6. ¿Cuál es su relación con la zona?", orden)

    escribir_lista(ws, "I", 8, frec)

    # -----------------------------------
    # 5 EDAD
    # -----------------------------------

    orden = [
        "18_a_29_anos",
        "30_a_44_anos",
        "45_a_64_anos",
        "65_anos_o_mas",
        "Vacio"
    ]

    frec = contar_frecuencias(
        df,
        "3. Edad (en años cumplidos): marque una categoría que incluya su edad.",
        orden
    )

    escribir_lista(ws, "D", 29, frec)

    # -----------------------------------
    # 6 ESCOLARIDAD
    # -----------------------------------

    orden = [
        "ninguna",
        "primaria_completa",
        "primaria_incompleta",
        "secundaria_completa",
        "secundaria_incompleta",
        "tecnico",
        "universitaria_completa",
        "universitaria_incompleta"
    ]

    frec = contar_frecuencias(df, "5. Escolaridad:", orden)

    escribir_lista(ws, "D", 39, frec)

    # -----------------------------------
    # 7 GENERO
    # -----------------------------------

    orden = [
        "masculino",
        "femenino",
        "persona_no_binaria"
    ]

    frec = contar_frecuencias(
        df,
        "4. ¿Con cuál de estas opciones se identifica?",
        orden
    )

    escribir_lista(ws, "D", 52, frec)

    # -----------------------------------
    # 8 SEGURIDAD DISTRITO
    # -----------------------------------

    orden = [
        "muy_inseguro",
        "inseguro",
        "ni_seguro_ni_inseguro",
        "seguro",
        "muy_seguro"
    ]

    frec = contar_frecuencias(
        df,
        "7. ¿Qué tan seguro percibe usted el distrito donde reside o transita?",
        orden
    )

    escribir_lista(ws, "C", 283, frec)

    # -----------------------------------
    # 9 CAMBIO SEGURIDAD
    # -----------------------------------

    orden = [
        "mucho_menos_seguro_1",
        "menos_seguro_2",
        "se_mantiene_igual_3",
        "mas_seguro_4",
        "mucho_mas_seguro_5"
    ]

    frec = contar_frecuencias(
        df,
        "8. En comparación con los 12 meses anteriores, ¿cómo percibe que ha cambiado la seguridad en este distrito?",
        orden
    )

    escribir_lista(ws, "C", 291, frec)

    # -----------------------------------
    # 10 SEGURIDAD LUGARES
    # -----------------------------------

    columnas = [
        "seg_discotecas_bares",
        "seg_espacios_recreativos",
        "seg_lugar_residencia",
        "seg_paradas_estaciones",
        "seg_puentes_peatonales",
        "seg_transporte_publico",
        "seg_zona_bancaria",
        "seg_zona_comercio",
        "seg_zonas_residenciales",
        "seg_zonas_francas",
        "seg_lugares_turisticos",
        "seg_centros_educativos"
    ]

    orden = [
        "muy_inseguro_1",
        "inseguro_2",
        "ni_seguro_ni_inseguro_3",
        "seguro_4",
        "muy_seguro_5",
        "no_aplica"
    ]

    columnas_destino = ["B", "D", "F", "H", "J", "L"]

    for opcion, col_excel in zip(orden, columnas_destino):

        resultados = []

        for c in columnas:
            resultados.append((df[c] == opcion).sum())

        escribir_lista(ws, col_excel, 300, resultados)

    # -----------------------------------
    # 11 VICTIMIZACION
    # -----------------------------------

    orden = [
        "no",
        "si_pero_no_denuncie",
        "si_y_denuncie"
    ]

    frec = contar_frecuencias(
        df,
        "30. Durante los últimos 12 meses, ¿usted o algún miembro de su hogar fue afectado por algún delito?",
        orden
    )

    escribir_lista(ws, "D", 314, frec)

    # -----------------------------------
    # 12 MOTIVOS NO DENUNCIA
    # -----------------------------------
    
    orden = [
        "distancia_o_dificultad_de_acceso_a_oficinas_para_denunciar",
        "miedo_a_represalias",
        "falta_de_respuesta_o_seguimiento_en_denuncias_anteriores",
        "complejidad_al_colocar_la_denuncia",
        "desconocimiento_de_donde_colocar_la_denuncia",
        "el_policia_me_dijo_que_era_mejor_no_denunciar",
        "falta_de_tiempo_para_colocar_la_denuncia",
        "desconfianza_en_las_autoridades_o_en_el_proceso_de_denuncia",
        "otro_motivo"
    ]
    
    col = "30.2 En caso de NO haber realizado la denuncia, indique ¿cuál o cuáles fueron el motivo?"
    
    serie = df[col].dropna().astype(str)
    
    # separo las respuestas
    todas = []
    
    for valor in serie:
        partes = [x.strip().lower() for x in valor.split(",")]
        todas.extend(partes)
    
    # contar frecuencias
    conteo = pd.Series(todas).value_counts()
    
    # construir lista final en orden
    frec = [int(conteo.get(opcion, 0)) for opcion in orden]
    
    escribir_lista(ws, "D", 322, frec)

    # -----------------------------------
    # 13 HORARIO DELITO
    # -----------------------------------
    
    orden = [
        "00_00_02_59_madrugada",
        "03_00_05_59_madrugada",
        "06_00_08_59_manana",
        "09_00_11_59_manana",
        "12_00_14_59_mediodia_tarde",
        "15_00_17_59_tarde",
        "18_00_20_59_noche",
        "21_00_23_59_noche",
        "desconocido"
    ]
    
    col = "30.3 ¿Tiene conocimiento sobre el horario en el cual se presentó el hecho o situación que le afectó a usted o un familiar?"
    
    serie = df[col].dropna().astype(str)
    
    # separar todas las respuestas
    todas = []
    
    for valor in serie:
        partes = [x.strip().lower() for x in valor.split(",")]
        todas.extend(partes)
    
    # contar frecuencias
    conteo = pd.Series(todas).value_counts()
    
    # construir lista final
    frec = [int(conteo.get(opcion, 0)) for opcion in orden]
    
    escribir_lista(ws, "D", 336, frec)

    # -----------------------------------
    # 14 METODOLOGIA DELITO
    # -----------------------------------
    
    orden = [
        "arma_blanca_cuchillo_machete_tijeras",
        "arma_de_fuego",
        "amenazas_o_intimidacion",
        "arrebato_le_quitaron_un_objeto_de_forma_rapida_o_sorpresiva",
        "boquete_ingreso_mediante_apertura_de_huecos_en_paredes_techos_o_estructuras",
        "ganzua_pata_de_chancho_llaves_falsas_u_objetos_similares",
        "engano_mediante_mentiras_falsas_ofertas_o_distraccion",
        "escalamiento_ingreso_trepando_muros_rejas_o_techos",
        "no_sabe_no_recuerda",
        "otro"
    ]
    
    col = "30.4 ¿Cuál fue la forma o modo en que ocurrió la situación que afectó a usted o a algún miembro de su hogar?"
    
    serie = df[col].dropna().astype(str)
    
    # separar todas las respuestas
    todas = []
    
    for valor in serie:
        partes = [x.strip().lower() for x in valor.split(",")]
        todas.extend(partes)
    
    # contar frecuencias
    conteo = pd.Series(todas).value_counts()
    
    # construir lista final
    frec = [int(conteo.get(opcion, 0)) for opcion in orden]
    
    escribir_lista(ws, "D", 350, frec)

    # -----------------------------------
    # 15 CONFIANZA EN FUERZA PUBLICA
    # -----------------------------------
    
    col = [c for c in df.columns if "nivel de confianza" in c.lower()][0]
    
    # convertir a número (por si viene como texto)
    serie = pd.to_numeric(df[col], errors='coerce')
    
    # conteos por rango
    ninguna_confianza = ((serie >= 1) & (serie <= 2)).sum()
    poca_confianza = ((serie >= 3) & (serie <= 4)).sum()
    confiable = ((serie >= 5) & (serie <= 6)).sum()
    confianza_razonable = ((serie >= 7) & (serie <= 8)).sum()
    mucha_confianza = ((serie >= 9) & (serie <= 10)).sum()
    
    frec = [
        int(ninguna_confianza),
        int(poca_confianza),
        int(confiable),
        int(confianza_razonable),
        int(mucha_confianza)
    ]
    
    escribir_lista(ws, "D", 363, frec)

    # -----------------------------------
    # 16 PROFESIONALIDAD FUERZA PUBLICA
    # -----------------------------------
    
    col = [c for c in df.columns if "profesionalidad de la fuerza pública" in c.lower()][0]
    
    # convertir a número
    serie = pd.to_numeric(df[col], errors='coerce')
    
    # rangos
    nada_profesional = ((serie >= 1) & (serie <= 3)).sum()
    profesional = ((serie >= 4) & (serie <= 7)).sum()
    muy_profesional = ((serie >= 8) & (serie <= 10)).sum()
    
    frec = [
        int(nada_profesional),
        int(profesional),
        int(muy_profesional)
    ]
    
    escribir_lista(ws, "D", 373, frec)

    # -----------------------------------
    # 17 IDENTIFICACION POLICIAS
    # -----------------------------------
    
    orden = [
        "no",
        "si"
    ]
    
    frec = contar_frecuencias(
        df,
        "31. ¿Identifica usted a los policías de la Fuerza Pública de Costa Rica en su comunidad?",
        orden
    )
    
    escribir_lista(ws, "D", 381, frec)

   # -----------------------------------
    # 18 INTERACCION CON POLICIA
    # -----------------------------------
    
    orden = [
        "solicitud_de_ayuda_o_auxilio",
        "atencion_relacionada_con_una_denuncia",
        "atencion_cordial_o_preventiva_durante_un_patrullaje",
        "fui_abordado_o_registrado_para_identificacion",
        "fui_objeto_de_una_infraccion_o_conflicto",
        "evento_preventivos_civico_policial_reunion_comunitaria",
        "otra_(especifique)"
    ]
    
    col = "31.1 ¿Cuáles de los siguientes tipos de atención ha tenido?"
    
    serie = df[col].dropna().astype(str)
    
    # separar respuestas múltiples
    todas = []
    
    for valor in serie:
        partes = [x.strip().lower() for x in valor.split(",")]
        todas.extend(partes)
    
    # contar frecuencias
    conteo = pd.Series(todas).value_counts()
    
    # construir lista final en orden (incluye manejo de "otro")
    frec = []
    
    for opcion in orden:
        if opcion == "otra_(especifique)":
            valor = sum(conteo.get(x, 0) for x in conteo.index if "otra" in x)
        else:
            valor = conteo.get(opcion, 0)
    
        frec.append(int(valor))
    
    escribir_lista(ws, "D", 386, frec)
#___________________FINAL__________________

    archivo = io.BytesIO()
    wb.save(archivo)
    archivo.seek(0)
    
    return archivo
