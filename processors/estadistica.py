from openpyxl import load_workbook


def procesar_estadistica(ruta_estadistica, ruta_info_engine):
    # cargar archivos
    wb_est = load_workbook(ruta_estadistica, data_only=True)
    ws_est = wb_est.active  # hoja1

    wb_info = load_workbook(ruta_info_engine)
    ws_info = wb_info.active  # hoja1

    # función reutilizable para copiar rangos
    def copiar_rango(ws_origen, ws_destino, fila_ini_o, col_ini_o, fila_fin_o, col_fin_o,
                     fila_ini_d, col_ini_d):
        
        for i, fila in enumerate(range(fila_ini_o, fila_fin_o + 1)):
            for j, col in enumerate(range(col_ini_o, col_fin_o + 1)):
                valor = ws_origen.cell(row=fila, column=col).value
                ws_destino.cell(
                    row=fila_ini_d + i,
                    column=col_ini_d + j,
                    value=valor
                )

    # -----------------------------
    # 1. denuncias_distrito
    # A4:B16 → A165:B177
    copiar_rango(ws_est, ws_info, 4, 1, 16, 2, 165, 1)

    # -----------------------------
    # 2. horarios
    # F18:Q27 → F179:Q188
    copiar_rango(ws_est, ws_info, 18, 6, 27, 17, 179, 6)

    # -----------------------------
    # 3. modalidad
    # D34:O43 → D195:O204
    copiar_rango(ws_est, ws_info, 34, 4, 43, 15, 195, 4)

    # -----------------------------
    # 4. dias
    # D48:O55 → D222:O229
    copiar_rango(ws_est, ws_info, 48, 4, 55, 15, 222, 4)

    # guardar cambios
    wb_info.save(ruta_info_engine)
